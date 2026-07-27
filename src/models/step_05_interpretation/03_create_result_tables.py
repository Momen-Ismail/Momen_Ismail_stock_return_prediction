"""Create report-ready CSV tables and one formatted Excel workbook."""

from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(PROJECT_ROOT))

from src.config import (  # noqa: E402
    INTERPRETATION_OUTPUT_DIR,
    MODEL_OUTPUT_DIR,
)


OPTIMIZATION_DIR = MODEL_OUTPUT_DIR / "optimization"
TUNING_DIR = MODEL_OUTPUT_DIR / "tuning"

TUNED_FAMILIES = [
    "pls",
    "elastic_net",
    "random_forest",
    "gradient_boosting",
]

DISPLAY_NAMES = {
    "historical_mean": "Historical Mean",
    "ols_3": "OLS-3",
    "pls": "PLS",
    "elastic_net": "Elastic Net",
    "random_forest": "Random Forest",
    "gradient_boosting": "Gradient Boosting",
}


def read_csv(path, required_columns):
    """Read one required CSV and check its columns."""
    if not path.exists():
        raise FileNotFoundError(
            f"Missing required file: {path}"
        )

    data = pd.read_csv(path)

    missing = set(required_columns) - set(data.columns)

    if missing:
        raise ValueError(
            f"{path.name} is missing columns: "
            f"{sorted(missing)}"
        )

    return data


def add_display_names(data):
    """Add readable names for report tables."""
    result = data.copy()

    result["display_name"] = (
        result["model"]
        .map(DISPLAY_NAMES)
        .fillna(result["model"])
    )

    return result


def final_results():
    """Load final test model results."""
    path = (
        INTERPRETATION_OUTPUT_DIR
        / "model_interpretation_metrics.csv"
    )

    data = read_csv(
        path,
        [
            "rank",
            "model",
            "monthly_mse",
            "monthly_rmse",
            "monthly_oos_r2",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
            "prediction_bias",
        ],
    )

    data = add_display_names(data)

    return data[
        [
            "rank",
            "model",
            "display_name",
            "monthly_mse",
            "monthly_rmse",
            "monthly_oos_r2",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
            "prediction_bias",
        ]
    ].sort_values("rank").reset_index(drop=True)


def behavior_flags():
    """Load final model-behavior indicators."""
    path = (
        INTERPRETATION_OUTPUT_DIR
        / "model_behavior_flags.csv"
    )

    data = read_csv(
        path,
        [
            "model",
            "beats_historical_mean",
            "positive_monthly_oos_r2",
            "positive_pooled_oos_r2",
            "constant_or_near_constant_predictions",
            "positive_in_most_test_years",
            "best_prediction_model",
            "worst_prediction_model",
        ],
    )

    return add_display_names(data)


def fixed_vs_optimized():
    """Load the validation comparison created in Step 3."""
    path = (
        OPTIMIZATION_DIR
        / "fixed_vs_optimized_family_summary.csv"
    )

    return read_csv(
        path,
        [
            "model_family",
            "fixed_model",
            "optimized_model",
            "fixed_monthly_mse",
            "optimized_monthly_mse",
            "monthly_mse_improvement",
            "monthly_mse_improvement_percent",
            "fixed_monthly_oos_r2",
            "optimized_monthly_oos_r2",
            "optimization_improved",
        ],
    ).sort_values(
        "optimized_monthly_mse"
    ).reset_index(drop=True)


def best_hyperparameters():
    """Collect the selected tuning parameters."""
    tables = []

    for family in TUNED_FAMILIES:
        path = (
            TUNING_DIR
            / f"{family}_best_parameters.csv"
        )

        data = read_csv(
            path,
            [
                "model_family",
                "parameters",
            ],
        )

        if len(data) != 1:
            raise ValueError(
                f"{path.name} should contain exactly one row."
            )

        tables.append(data)

    return pd.concat(
        tables,
        ignore_index=True,
    )


def yearly_results():
    """Load annual test-period results."""
    path = (
        INTERPRETATION_OUTPUT_DIR
        / "yearly_prediction_results.csv"
    )

    data = read_csv(
        path,
        [
            "rank",
            "year",
            "model",
            "observations",
            "months",
            "monthly_mse",
            "monthly_rmse",
            "monthly_oos_r2",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
        ],
    )

    data = add_display_names(data)

    return data.sort_values(
        [
            "year",
            "rank",
        ]
    ).reset_index(drop=True)


def interpretation_table(filename, required_columns):
    """Load one table produced by interpretation file 02."""
    path = (
        INTERPRETATION_OUTPUT_DIR
        / filename
    )

    return read_csv(
        path,
        required_columns,
    )


def predictor_groups():
    """Combine linear and tree predictor-family summaries."""
    linear = interpretation_table(
        "linear_coefficients_by_predictor_group.csv",
        [
            "model",
            "predictor_group",
            "total_absolute_coefficient",
            "average_absolute_coefficient",
            "predictor_count",
            "nonzero_predictor_count",
            "nonzero_predictor_share",
        ],
    ).rename(
        columns={
            "total_absolute_coefficient": "total_importance",
            "average_absolute_coefficient": "average_importance",
        }
    )

    linear["importance_type"] = (
        "absolute_standardized_coefficient"
    )

    tree = interpretation_table(
        "tree_importance_by_predictor_group.csv",
        [
            "model",
            "predictor_group",
            "total_importance",
            "average_importance",
            "predictor_count",
            "nonzero_predictor_count",
            "nonzero_predictor_share",
        ],
    )

    tree["importance_type"] = (
        "impurity_importance"
    )

    result = pd.concat(
        [
            linear,
            tree,
        ],
        ignore_index=True,
    )

    result = add_display_names(result)

    return result[
        [
            "model",
            "display_name",
            "importance_type",
            "predictor_group",
            "total_importance",
            "average_importance",
            "predictor_count",
            "nonzero_predictor_count",
            "nonzero_predictor_share",
        ]
    ].sort_values(
        [
            "model",
            "total_importance",
        ],
        ascending=[
            True,
            False,
        ],
    ).reset_index(drop=True)


def format_sheet(sheet, data):
    """Apply simple formatting to one Excel worksheet."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E2F3",
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for column_number, column_name in enumerate(
        data.columns,
        start=1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        values = [
            str(column_name),
            *[
                ""
                if pd.isna(value)
                else str(value)
                for value in data[column_name].head(200)
            ],
        ]

        width = min(
            max(len(value) for value in values) + 2,
            42,
        )

        sheet.column_dimensions[
            column_letter
        ].width = width

        if pd.api.types.is_float_dtype(
            data[column_name]
        ):
            for cell in sheet[column_letter][1:]:
                if column_name.endswith("_share"):
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "0.000000"


def write_workbook(tables, path):
    """Write all final tables into one workbook."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, data in tables.items():
        if data.empty:
            raise ValueError(
                f"{sheet_name} is empty."
            )

        sheet = workbook.create_sheet(
            sheet_name
        )

        for row in dataframe_to_rows(
            data,
            index=False,
            header=True,
        ):
            sheet.append(row)

        format_sheet(
            sheet,
            data,
        )

    workbook.save(path)


def main():
    """Create final CSV tables and Excel workbook."""
    INTERPRETATION_OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    final = final_results()
    flags = behavior_flags()
    comparison = fixed_vs_optimized()
    parameters = best_hyperparameters()
    yearly = yearly_results()

    complexity = interpretation_table(
        "model_complexity_summary.csv",
        [
            "model",
            "number_of_predictors",
        ],
    )

    top_features = interpretation_table(
        "feature_importance_top_variables.csv",
        [
            "model",
            "importance_type",
            "predictor",
            "predictor_group",
            "signed_value",
            "absolute_value",
            "rank",
        ],
    )

    specifications = interpretation_table(
        "final_model_specifications.csv",
        [
            "model",
            "parameters",
        ],
    )

    groups = predictor_groups()

    final.to_csv(
        INTERPRETATION_OUTPUT_DIR
        / "final_prediction_results.csv",
        index=False,
    )

    comparison.to_csv(
        INTERPRETATION_OUTPUT_DIR
        / "fixed_vs_optimized_results.csv",
        index=False,
    )

    parameters.to_csv(
        INTERPRETATION_OUTPUT_DIR
        / "best_hyperparameters.csv",
        index=False,
    )

    groups.to_csv(
        INTERPRETATION_OUTPUT_DIR
        / "final_predictor_group_summary.csv",
        index=False,
    )

    workbook_path = (
        INTERPRETATION_OUTPUT_DIR
        / "final_report_results.xlsx"
    )

    write_workbook(
        {
            "Final Results": final,
            "Behavior Flags": flags,
            "Fixed vs Optimized": comparison,
            "Hyperparameters": parameters,
            "Yearly Results": yearly,
            "Model Complexity": complexity,
            "Top Features": top_features,
            "Predictor Groups": groups,
            "Specifications": specifications,
        },
        workbook_path,
    )

    print("\nSaved final report workbook:")
    print(workbook_path)


if __name__ == "__main__":
    main()