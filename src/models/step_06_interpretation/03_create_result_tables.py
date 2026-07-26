"""Create report-ready CSV tables and a formatted Excel workbook."""

from ast import literal_eval
from importlib import import_module
import json
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


PROJECT_ROOT = Path(__file__).resolve().parents[3]
sys.path.append(str(PROJECT_ROOT))

utils = import_module(
    "src.models.step_06_interpretation.00_utils"
)


TUNED_FAMILIES = [
    "pls",
    "elastic_net",
    "decision_tree",
    "random_forest",
]


def final_results():
    """Load the final-test ranking created in Step 6 file 01."""
    path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "model_interpretation_metrics.csv"
    )

    data = utils.read_csv(
        path,
        "Step 6 file 01",
        [
            "rank",
            "model",
            "monthly_mse",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
        ],
    )

    data = data[
        [
            "rank",
            "model",
            "monthly_mse",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
        ]
    ].sort_values(
        "rank"
    ).reset_index(drop=True)

    utils.ensure_unique(
        data,
        ["model"],
        "Final prediction results",
    )

    return data


def fixed_vs_optimized():
    """Compare fixed and optimized validation results."""
    path = (
        utils.OPTIMIZATION_DIR
        / "fixed_vs_optimized_all_metrics.csv"
    )

    data = utils.read_csv(
        path,
        "Step 3 fixed-versus-optimized comparison",
        [
            "model",
            "sample",
            "monthly_mse",
            "pooled_rmse",
            "pooled_mae",
            "oos_r2",
            "prediction_target_correlation",
        ],
    )

    data = data[
        data["sample"].eq("validation")
    ].copy()

    rows = []

    for family in TUNED_FAMILIES:
        fixed_name = f"{family}_fixed"
        optimized_name = f"{family}_optimized"

        fixed = data[
            data["model"].eq(fixed_name)
        ]

        optimized = data[
            data["model"].eq(optimized_name)
        ]

        if len(fixed) != 1:
            raise ValueError(
                "Expected exactly one validation row "
                f"for {fixed_name}, found {len(fixed)}."
            )

        if len(optimized) != 1:
            raise ValueError(
                "Expected exactly one validation row "
                f"for {optimized_name}, "
                f"found {len(optimized)}."
            )

        fixed = fixed.iloc[0]
        optimized = optimized.iloc[0]

        rows.append({
            "model_family": family,
            "fixed_model": fixed_name,
            "optimized_model": optimized_name,

            "fixed_monthly_mse": (
                fixed["monthly_mse"]
            ),
            "optimized_monthly_mse": (
                optimized["monthly_mse"]
            ),
            "monthly_mse_improvement": (
                fixed["monthly_mse"]
                - optimized["monthly_mse"]
            ),

            "fixed_rmse": (
                fixed["pooled_rmse"]
            ),
            "optimized_rmse": (
                optimized["pooled_rmse"]
            ),
            "rmse_improvement": (
                fixed["pooled_rmse"]
                - optimized["pooled_rmse"]
            ),

            "fixed_mae": (
                fixed["pooled_mae"]
            ),
            "optimized_mae": (
                optimized["pooled_mae"]
            ),
            "mae_improvement": (
                fixed["pooled_mae"]
                - optimized["pooled_mae"]
            ),

            "fixed_oos_r2": (
                fixed["oos_r2"]
            ),
            "optimized_oos_r2": (
                optimized["oos_r2"]
            ),
            "oos_r2_change": (
                optimized["oos_r2"]
                - fixed["oos_r2"]
            ),

            "fixed_correlation": (
                fixed[
                    "prediction_target_correlation"
                ]
            ),
            "optimized_correlation": (
                optimized[
                    "prediction_target_correlation"
                ]
            ),
            "correlation_change": (
                optimized[
                    "prediction_target_correlation"
                ]
                - fixed[
                    "prediction_target_correlation"
                ]
            ),
        })

    result = (
        pd.DataFrame(rows)
        .sort_values("optimized_monthly_mse")
        .reset_index(drop=True)
    )

    utils.ensure_unique(
        result,
        ["model_family"],
        "Fixed versus optimized table",
    )

    return result


def best_hyperparameters():
    """Load the selected tuning parameters without rebuilding models."""
    rows = []

    for family in TUNED_FAMILIES:
        path = (
            utils.TUNING_DIR
            / f"{family}_best_parameters.csv"
        )

        data = utils.read_csv(
            path,
            "Step 2 model tuning",
            [
                "model_family",
                "parameters",
            ],
        )

        if len(data) != 1:
            raise ValueError(
                f"Expected one selected parameter row for "
                f"{family}, found {len(data)}."
            )

        parameters = literal_eval(
            data.iloc[0]["parameters"]
        )

        if not isinstance(parameters, dict):
            raise ValueError(
                f"Selected parameters for {family} "
                "must be stored as a dictionary."
            )

        row = {
            "model": family,
            **parameters,
            "parameters_json": json.dumps(
                parameters,
                sort_keys=True,
            ),
        }

        rows.append(row)

    result = pd.DataFrame(rows)

    utils.ensure_unique(
        result,
        ["model"],
        "Best hyperparameters",
    )

    return result


def yearly_results():
    """Load yearly final-test results created in Step 6 file 01."""
    path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "yearly_prediction_results.csv"
    )

    data = utils.read_csv(
        path,
        "Step 6 file 01",
        [
            "rank",
            "year",
            "model",
            "observations",
            "monthly_mse",
            "rmse",
            "mae",
            "oos_r2",
            "correlation",
        ],
    )

    data = data.sort_values(
        [
            "year",
            "rank",
        ]
    ).reset_index(drop=True)

    utils.ensure_unique(
        data,
        [
            "model",
            "year",
        ],
        "Yearly prediction results",
    )

    return data


def complexity_summary():
    """Load fitted complexity measures saved from the Step 5 models."""
    path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "model_complexity_summary.csv"
    )

    fitted = utils.read_csv(
        path,
        "Step 6 file 02",
        [
            "model",
            "number_of_predictors",
        ],
    )

    historical_mean = pd.DataFrame([
        {
            "model": "historical_mean",
            "number_of_predictors": 0,
        }
    ])

    result = pd.concat(
        [
            historical_mean,
            fitted,
        ],
        ignore_index=True,
        sort=False,
    )

    model_order = {
        model: position
        for position, model in enumerate(
            utils.ACTIVE_MODELS
        )
    }

    result["_model_order"] = (
        result["model"]
        .map(model_order)
    )

    result = (
        result
        .sort_values("_model_order")
        .drop(columns="_model_order")
        .reset_index(drop=True)
    )

    utils.ensure_unique(
        result,
        ["model"],
        "Model complexity summary",
    )

    return result


def top_features():
    """Load the unified top-feature table from Step 6 file 02."""
    path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "feature_importance_top_variables.csv"
    )

    data = utils.read_csv(
        path,
        "Step 6 file 02",
        [
            "model",
            "importance_type",
            "predictor",
            "signed_value",
            "absolute_value",
            "rank",
        ],
    )

    data = data.sort_values(
        [
            "model",
            "importance_type",
            "rank",
        ]
    ).reset_index(drop=True)

    utils.ensure_unique(
        data,
        [
            "model",
            "importance_type",
            "rank",
        ],
        "Top features",
    )

    return data


def behavior_flags():
    """Load model-behavior indicators created in Step 6 file 01."""
    path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "model_behavior_flags.csv"
    )

    data = utils.read_csv(
        path,
        "Step 6 file 01",
        [
            "model",
            "beats_historical_mean",
            "positive_oos_r2",
            "constant_or_near_constant_predictions",
            "positive_in_most_test_years",
            "best_prediction_model",
            "worst_prediction_model",
        ],
    )

    utils.ensure_unique(
        data,
        ["model"],
        "Model behavior flags",
    )

    return data


def format_worksheet(sheet, data):
    """Apply simple academic formatting to one worksheet."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E2F3",
    )

    thin_fill = PatternFill(
        fill_type="solid",
        fgColor="F3F6FA",
    )

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.font = Font(
            bold=True
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for row_number in range(
        2,
        sheet.max_row + 1,
    ):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = thin_fill

    for column_number, column_name in enumerate(
        data.columns,
        start=1,
    ):
        column_letter = get_column_letter(
            column_number
        )

        values = [
            str(column_name)
        ] + [
            ""
            if pd.isna(value)
            else str(value)
            for value in data[column_name].head(200)
        ]

        width = min(
            max(
                len(value)
                for value in values
            ) + 2,
            38,
        )

        sheet.column_dimensions[
            column_letter
        ].width = width

        if pd.api.types.is_integer_dtype(
            data[column_name]
        ):
            number_format = "0"

        elif pd.api.types.is_float_dtype(
            data[column_name]
        ):
            number_format = "0.000000"

        else:
            number_format = None

        if number_format is not None:
            for cell in sheet[
                column_letter
            ][1:]:
                cell.number_format = (
                    number_format
                )

        for cell in sheet[
            column_letter
        ][1:]:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=False,
            )


def write_workbook(tables, path):
    """Write all report tables to one formatted workbook."""
    workbook = Workbook()

    workbook.remove(
        workbook.active
    )

    for sheet_name, data in tables.items():
        sheet = workbook.create_sheet(
            sheet_name
        )

        for row in dataframe_to_rows(
            data,
            index=False,
            header=True,
        ):
            sheet.append(row)

        format_worksheet(
            sheet,
            data,
        )

    workbook.save(path)


def main():
    """Create all report-ready tables."""
    final = final_results()
    comparison = fixed_vs_optimized()
    parameters = best_hyperparameters()
    yearly = yearly_results()
    complexity = complexity_summary()
    features = top_features()
    flags = behavior_flags()

    utils.save_csv(
        final,
        "final_prediction_results.csv",
    )

    utils.save_csv(
        comparison,
        "fixed_vs_optimized_results.csv",
    )

    utils.save_csv(
        parameters,
        "best_hyperparameters.csv",
    )

    # Keep the raw model-complexity file from file 02 unchanged.
    utils.save_csv(
        complexity,
        "final_model_complexity.csv",
    )

    workbook_path = (
        utils.INTERPRETATION_OUTPUT_DIR
        / "final_report_results.xlsx"
    )

    write_workbook(
        {
            "Final Results": final,
            "Behavior Flags": flags,
            "Fixed vs Optimized": comparison,
            "Hyperparameters": parameters,
            "Yearly Results": yearly,
            "Model Complexity": complexity,
            "Top Features": features,
        },
        workbook_path,
    )

    print(
        "Saved report tables to:"
    )

    print(
        utils.INTERPRETATION_OUTPUT_DIR
    )

    print(
        "\nSaved Excel workbook:"
    )

    print(
        workbook_path
    )


if __name__ == "__main__":
    main()